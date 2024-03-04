import pandas as pd
from openpyxl import load_workbook
import sys
import subprocess
import py7zr
import time
import os

def extract():
    with py7zr.SevenZipFile(source_7z,mode="r",password=password) as file: 
        file.extractall() #extracts the .7z file
    return ".\\excel_file.xlsx"

def compress():
    os.rename(source_7z,source_7z+"temp") #rename the compressed archive
    try:
        # Compress the modified content into a new encrypted archive
        with py7zr.SevenZipFile(source_7z, mode="w", password=password) as archive:
            archive.writeall("extracted-folder")
    except Exception as e:
        # If compression fails, restore the original archive
        os.rename(source_7z+"temp",source_7z)
        raise e
    else:
        # If compression succeeds, delete the temporary archive
        os.remove(source_7z)
    
#global members
os.chdir("directory-of-.7z file") #can be skipped if the archive is in the cwd
source_7z="your-.7zfile"
password="your-password"
source=extract() 
obj=list() #obj is a list of objects
software_list_without_duplicates=list()

#class node stores the software name and its prefix number(index), username and its prefix number(index2) and the password
class node: 
    def __init__(self,s,i,u,i2,p):
        self.software=s
        self.index=i
        self.username=u
        self.index2=i2
        self.password=p
        
    # def display(self):
    #     print(f"{self.index}. SOFTWARE : {self.software}\n{self.index2}. USERNAME : {self.username}\nPASSWORD: {self.password}")
    # # for testing purposes only
    
def main():
    data=pd.read_excel(source,sheet_name=0,usecols=[0,2,4])
    data=data.dropna() #remove nan values
    #after removing nan values, software,username and password are stored in 0th,1st and 2nd column respectively, change it based on ur requirements
    global software_list
    software_list=data.iloc[:,0].tolist() #storing all rows of 0th column of the excel file into a list called software
    print(software_list)
    for i in software_list:
        if i not in software_list_without_duplicates:
            software_list_without_duplicates.append(i) #removing duppicate values from software list to obtain software prefix number
    username_list=data.iloc[:,1].tolist() #storing all rows of 1st column of the excel file into a list called username
    password_list=data.iloc[:,2].tolist() #storing all rows of 2nd column of the excel file into a list called password
    c1=1 #software prefix number
    for m in software_list_without_duplicates:
        c2=1 #username prefix number
        for i,j,k in zip(software_list,username_list,password_list): #this line iterates through all the three lists at the same time; i,j,k are incremented simultaneousy
            if i!=m:
                continue 
            obj.append(node(i,c1,j,c2,k))
            c2+=1
        c1+=1
        
    # for i in obj:
    #     i.display()
    # # for testing purposes only
    
    while True:
        print(r'''
        /=================================OPERATIONS=====================================\
        |1. ADD A PASSWORD                                                               |
        |2. EXTRACT A PASSWORD                                                           |
        |3. DELETE A PASSWORD                                                            |
        |4. OPEN SOURCE                                                                  |
        |5. SAVE                                                                         |
        |6. EXIT                                                                         | 
        \================================================================================/''')
        ans=int(input("\nEnter your choice: "))
        print("")
        if ans==1:
            add()
        elif ans==2:
            display()
        elif ans==3:
            delete()
        elif ans==4:
            #find the excel.exe file in ur system, the below path worked for my windows 11system
            subprocess.Popen(["C:\\Program Files\\Microsoft Office\\root\\Office16\\EXCEL.EXE",'/r',source]) #runs excel.exe and opens microsoft excel in read-only mode 
            time.sleep(10) #user has 10 seconds to view the file
            os.system("taskkill /f /im EXCEL.exe") #closes microsoft excel
        elif ans==5:
            compress()
            os.system(f"rmdir /S /Q {"extracted-folder"}") #removes the previously extracted file
            extract()
            print("SAVE SUCCESSFUL")
        elif ans==6:
            compress()
            os.system(f"rmdir /S /Q {"extracted-folder"}") #removes the previously extracted file
            sys.exit()
        else:
            print("INVALID")
                    
def display():
    while True:
        m=0
        for i in range(len(software_list_without_duplicates)):
            print(f"{i+1} : {software_list_without_duplicates[i]}")
            m=i+1
        print(f"{m+1} : EXIT")
        ch=int(input("\nEnter your choice: "))
        if ch==m+1:
            return
        elif ch<=0 or ch>len(software_list_without_duplicates):
            print("\nINVALID\n")
            return
        print("")
        max=0
        for i in obj:
            if ch==i.index:
                print(f"{i.index2} : {i.username}")
                max=i.index2
        print(f"{max+1} : EXIT")
        ch2=int(input("\nEnter your choice: "))
        if ch2==max+1:
            return
        elif ch2<=0 or ch2>max+1:
            print("\nINVALID\n")
            return
        print("")
        for i in obj:
            if ch2==i.index2 and ch==i.index:
                print(f"PASSWORD : \033[92m {i.password}\033[00m\n") #ANSI code to display password in green color
        return

def add():
    wb=load_workbook(source)
    sheet=wb.active
    input_software=input("Enter Software (Enter 0 to Exit) : ").title()
    if input_software=='0':
        return
    input_username=input("Enter Username (Enter 0 to Exit) : ")
    if input_username=='0':
        return
    input_password=input("Enter Password (Enter 0 to Exit) : ")
    if input_password=='0':
        return
    software_list.append(input_software)
    if input_software not in software_list_without_duplicates:
        software_list_without_duplicates.append(input_software)
    row=3
    breaker=0
    i=0
    index,index2=1,1
    for i in range(len(obj)):
        if obj[i].software==input_software:
            index=obj[i].index
            index2=obj[i].index2
            try:
                while obj[i].software==input_software:
                    i+=1
                    row+=1
                    breaker=1
                    index2+=1
            except IndexError:
                i+=1
                breaker=1
                break
        if breaker==1:
            break
        row+=1
        index=obj[i].index+1
        index2=1
    if breaker==0:
        i+=1
    sheet.insert_rows(row) #adds a row to the excel file
    sheet['A'+str(row)].value,sheet['C'+str(row)].value,sheet['E'+str(row)].value=input_software,input_username,input_password #writes the values to the excel file
    obj.insert(i,node(input_software,index,input_username,index2,input_password)) #inserts a node to the list of objects
    wb.save(source)
    print("\nPASSWORD INSERTED\n")

def delete():
    wb=load_workbook(source)
    sheet=wb.active
    while True:
        m=0
        for i in range(len(software_list_without_duplicates)):
            print(f"{i+1} : {software_list_without_duplicates[i]}")
            m=i+1
        print(f"{m+1} : EXIT")
        ch=int(input("\nEnter your choice: "))
        if ch==m+1:
            return
        elif ch<=0 or ch>len(software_list_without_duplicates):
            print("\nINVALID\n")
            return
        print("")
        max=0
        for i in obj:
            if ch==i.index:
                print(f"{i.index2} : {i.username}")
                max=i.index2
        print(f"{max+1} : EXIT")
        ch2=int(input("\nEnter your choice: "))
        if ch2==max+1:
            return
        elif ch2<=0 or ch2>max+1:
            print("\nINVALID\n")
            return
        print("")
        c=[0 for _ in range(len(software_list_without_duplicates)+1)]
        p=0
        for m in software_list_without_duplicates:
            p+=1
            for n in software_list:
                if m==n:
                    c[p]+=1
        i=0
        row=3
        for i in range(len(obj)):
            if c[ch]!=1 and ch==obj[i].index and ch2==obj[i].index2:
                j=i+1
                try:
                    while obj[j].index==obj[i].index:
                        obj[j].index2-=1
                        j+=1
                        row+=1
                except IndexError:
                    pass
                break
            elif c[ch]==1 and ch==obj[i].index and ch2==1:
                j=i+1
                while j!=len(obj):
                    obj[j].index-=1
                    j+=1
                    row+=1
                break
            row+=1
        if c[ch]==1:
            software_list_without_duplicates.pop(ch-1) #pop the software name from the list
        else:
            software_list.remove(obj[i].software.title())
        obj.pop(i) #pop the object from the list
        sheet.delete_rows(row) #delete the row from the excel file
        wb.save(source)
        print("\nPASSWORD DELETED\n")
        break
            
main()